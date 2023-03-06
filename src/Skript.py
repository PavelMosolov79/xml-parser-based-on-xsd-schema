import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

title_list = ""
row = 1
peas_of_sheat = ""

def pars_root(element):
    global title_list, row
    res = {}
    if len(element)==0: return res

    for i in element:
        if 'element' in i.tag:
            for element in i[0]:
                if 'documentation' in element.tag:
                    title_list = element.text
                    print(element.text)


# Париснг схемы
def func(element):
    res = {}
    if len(element)==0: return res
    for i in element:
        if 'annotation' in i.tag:
            res['annotation']=i[0].text
        if 'sequence' in i.tag:
            for element in i:
                if 'element' in element.tag:
                    if 'name' in element.attrib.keys():
                        res[element.attrib['name']]=func(element)
        if 'complexType' in i.tag:
            for element in i[0]:
                if 'element' in element.tag:
                    if 'name' in element.attrib.keys():
                        res[element.attrib['name']]=func(element)
            
    return res

#Вывод xml файла на основе схемы xsd
def printXmlXsd(element, xsd, tabs=0):
    global row, cell

    # peas_of_sheat
    for i in element:
        tag = i.tag.split('}')[1]
        print("\n",tag, xsd.keys())
        # for t in range(tabs):
        # row+=1
        # peas_of_sheat['A'+str(row)] = ""
        # print('  ',end='')
        if tag in xsd.keys():
            if 'annotation' in xsd[tag].keys():
                row += 1
                fill_one = PatternFill('solid', fgColor="556B2F")
                fill_two = PatternFill('solid', fgColor="9ACD32")
                
                peas_of_sheat['A'+str(row)].fill = fill_one
                peas_of_sheat['B'+str(row)].fill = fill_two

                peas_of_sheat['A'+str(row)].border = Border(top=Side(border_style="thin", 
                    color="808000"))
                peas_of_sheat['B'+str(row)].border = Border(top=Side(border_style="thin", 
                    color="808000"), left=Side(border_style="thin", color="808000"))

                peas_of_sheat['A'+str(row)] = xsd[tag]['annotation']
                peas_of_sheat['B'+str(row)] = i.text if not i.text is None and len(i.text)>0 else 'NONE'
                print("КОЛОНКИ", xsd[tag]['annotation'], i.text if not i.text is None and len(i.text)>0 else '')
            else:
                # row += 1
                # peas_of_sheat['A'+str(row)] = ""
                # peas_of_sheat['B'+str(row)] = i.text if not i.text is None and len(i.text)>0 else 'NONE'
                print("ХЗ", tag, i.text if not i.text is None and len(i.text)>0 else '')
            printXmlXsd(i, xsd[tag], tabs+1)
        else:
            # row += 1
            # peas_of_sheat['A'+str(row)] = ""

            row += 1
            fill_one = PatternFill('solid', fgColor="90EE90")
            fill_two = PatternFill('solid', fgColor="98FB98")

            peas_of_sheat['A'+str(row)].fill = fill_one
            peas_of_sheat['B'+str(row)].fill = fill_two
            
            peas_of_sheat['A'+str(row)].border = Border(top=Side(border_style="thin", 
                color="808000"))
            peas_of_sheat['B'+str(row)].border = Border(top=Side(border_style="thin", 
                color="808000"), left=Side(border_style="thin", color="808000"))

            peas_of_sheat['A'+str(row)] = tag
            peas_of_sheat['B'+str(row)] = i.text if not i.text is None and len(i.text)>0 else 'NONE'
            print("123", tag, i.text if not i.text is None and len(i.text)>0 else '')
            printXmlXsd(i, {}, tabs+1)


def main():
    # Париснг схемы
    global peas_of_sheat
    xsd = ET.parse('forestDevelopmentProject.xsd')

    # Считывание файла xsd в root
    root = xsd.getroot()

    # Парсинг
    res={}




    for i in root:
        if 'complexType' in i.tag:
            # print(func(i))
            res[i.attrib['name']]=func(i)

    tree = ET.parse('forestDevelopmentProject_4_1.xml')
    r2 = tree.getroot()

    pars_root(root)

    #Создание нового файла excel
    new_file_excel = openpyxl.Workbook()
    new_file_excel.create_sheet(title = title_list, index = 0)
    
    peas_of_sheat = new_file_excel[title_list]
    peas_of_sheat.column_dimensions['A'].width = 100
    peas_of_sheat.column_dimensions['B'].width = 40

    peas_of_sheat['A1'].fill = PatternFill('solid', fgColor="556B2F")
    peas_of_sheat['A1'].font = Font(size= 23, underline='single', color='FFBB00', bold=True, italic=True)
    peas_of_sheat['A1'] = title_list

    printXmlXsd([r2], res)

    sav = "прикол"+".xlsx"
    new_file_excel.save(sav)



if __name__ == '__main__':
    main()



    #print("kkk",r2)

    # for i in root:
    #     if 'name' in i.attrib.keys():
    #         print(i.tag, i.attrib['name'])
    #     else:
    #         print(i.tag)

    # for i in r2:
    #     print(i.tag.split('}')[1])