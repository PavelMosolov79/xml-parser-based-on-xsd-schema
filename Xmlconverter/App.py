import sys 
import xml.etree.ElementTree as ET
import os
import openpyxl

from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

str_name = ""
str_text = ""

title_list = ""
row = 1
peas_of_sheat = ""

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.window_title()
        self.button_start()

    def window_title(self):
        self.setWindowTitle("File Decryptor")
        self.setFixedSize(QSize(700, 200))

    def button_start(self):
        #Кнопка для запуска декодирования файла
        button = QPushButton("Запуск декодирования", self)
        button.clicked.connect(self.button_start_clicked)

        #Кнопка обзор для xsd
        button_xsd = QPushButton("Обзор", self)
        button_xsd.clicked.connect(self.download_file_xsd)      

        #Кнопка обзор для xml
        button_xml = QPushButton("Обзор", self)
        button_xml.clicked.connect(self.download_file_xml)

        #Поле для ввода названия файла
        self.name_file = QTextEdit('Поле для ввода')
        # name_file.setText(')

        #Разделение на ячейки
        self.gridBox = QGridLayout()
        self.gridBox.setRowStretch(1, 20)
        self.gridBox.setRowStretch(2, 20)
        self.gridBox.setRowStretch(3, 20)
        # self.gridBox.setRowStretch(3, 200)
        self.lable_xsd = QLabel('Выберите файл xsd')
        self.lable_xml = QLabel('Выберите файл xml')
        self.lable_name = QLabel('Введите название файла в который будет сохранен результат:')

        self.gridBox.addWidget(button_xsd, 0, 1, alignment=Qt.AlignmentFlag.AlignRight)
        self.gridBox.addWidget(self.lable_xsd, 0, 0, alignment=Qt.AlignmentFlag.AlignCenter)
        self.gridBox.addWidget(button_xml, 1, 1, alignment=Qt.AlignmentFlag.AlignRight)
        self.gridBox.addWidget(self.lable_xml, 1, 0, alignment=Qt.AlignmentFlag.AlignCenter)
        self.gridBox.addWidget(self.name_file, 2, 1, alignment=Qt.AlignmentFlag.AlignRight)
        self.gridBox.addWidget(self.lable_name, 2, 0, alignment=Qt.AlignmentFlag.AlignCenter)

        self.gridBox.addWidget(button, 3, 1, alignment=Qt.AlignmentFlag.AlignRight)
        
        self.setLayout(self.gridBox)
    
    str_xsd = ""
    str_xml = ""
    str_file_name =""

    #первое поле для выбора файла xsd
    def download_file_xsd(self):
        self.str_xsd = QFileDialog.getOpenFileName(self, 'aboba', filter='xsd (*.xsd)')
        self.lable_xsd.setText(self.str_xsd[0])
        self.str_xsd = self.str_xsd[0]

    #Второе поле для выбора файла xml
    def download_file_xml(self):
        self.str_xml = QFileDialog.getOpenFileName(self, 'aboba', filter='xml (*.xml)')
        self.lable_xml.setText(self.str_xml[0])
        self.str_xml = self.str_xml[0]

    def name_file_func(self):
        self.str_file_name = self.name_file.toPlainText()
        return self.str_file_name

    def button_start_clicked(self):
        save_xsd = self.str_xsd
        save_xml = self.str_xml
        save_file_name = self.name_file_func()

        if (save_xsd == ""):
            MsgBox = QMessageBox()
            MsgBox.setWindowTitle("Ошибка!")
            MsgBox.setText("Файл xsd не был выбран!")
            MsgBox.exec()
        elif (save_xml == ""):
            MsgBox = QMessageBox()
            MsgBox.setWindowTitle("Ошибка!")
            MsgBox.setText("Файл xml не был выбран!")
            MsgBox.exec()
        elif (save_file_name == "" or save_file_name == "Поле для ввода"):
            MsgBox = QMessageBox()
            MsgBox.setWindowTitle("Ошибка!")
            MsgBox.setText("Вы не ввели название файла!")
            MsgBox.exec()
        else:
            ######################################################
                # Париснг схемы
                # Париснг схемы
            global peas_of_sheat
            global str_name
            print(save_xml, save_xsd)

            # Париснг схемы
            with open( save_xsd, 'rb+') as file:
                xsd = ET.parse(file)

                # Считывание файла xsd в root
                root = xsd.getroot()
                # print(root)

                # Парсинг
                res={}

                for i in root:
                    if 'complexType' in i.tag:
                        
                        res[i.attrib['name']]=func(i)

                tree = ET.parse(save_xml)
                r2 = tree.getroot()

                # print(r2)

                pars_root(root)
            # На всякий документ
            # str_name = save_file_name + str(".doc")

            
            # replace_element([r2])
            # printXmlXsd([r2], res)

            # file = open(str_name, 'w')
            # file.write(str_text)

            # file.close()

            new_file_excel = openpyxl.Workbook()
            new_file_excel.create_sheet(title = title_list, index = 0)
            
            peas_of_sheat = new_file_excel[title_list]
            peas_of_sheat.column_dimensions['A'].width = 100
            peas_of_sheat.column_dimensions['B'].width = 40

            peas_of_sheat['A1'].fill = PatternFill('solid', fgColor="556B2F")
            peas_of_sheat['A1'].font = Font(size= 23, underline='single', color='FFBB00', bold=True, italic=True)
            peas_of_sheat['A1'] = title_list

            #Создание нового файла excel
            printXmlXsd([r2], res)

            sav = save_file_name + ".xlsx"
            new_file_excel.save(sav)


            MsgBox = QMessageBox()
            MsgBox.setWindowTitle("Успех!")
            MsgBox.setText("Файл собран и находится в корневой дирректории")
            MsgBox.exec()


def replace_element(element):
    print(element)
    # element = element.replace('name', 'Имя:')
    # element = element.replace('date', 'Дата:')


def pars_root(element):
    global title_list, row
    res = {}
    if len(element)==0: return res

    for i in element:
        if 'element' in i.tag:
            for element in i[0]:
                if 'documentation' in element.tag:
                    title_list = element.text
                    print(element.text)
    # res = {}

    # if len(element)==0: return res

    # file = open(str_name, 'w')

    # for i in element:
    #     if 'element' in i.tag:
    #         for element in i[0]:
    #             if 'documentation' in element.tag:
    #                 file.write(element.text)
    #                 print(element.text)

    # file.close()

# Париснг схемы
def func(element):
    res = {}

    if len(element)==0: return res
    for i in element:
        if 'annotation' in i.tag:
            res['annotation']=i[0].text
        if 'sequence' in i.tag:
            for element in i:
                if 'element' in element.tag:
                    if 'name' in element.attrib.keys():
                        res[element.attrib['name']]=func(element)
        if 'complexType' in i.tag:
            for element in i[0]:
                if 'element' in element.tag:
                    if 'name' in element.attrib.keys():
                        res[element.attrib['name']]=func(element)
            
    return res


#Вывод xml файла на основе схемы xsd
def printXmlXsd(element, xsd, tabs=0):
    global peas_of_sheat
    global str_text
    global row, cell

    # peas_of_sheat
    for i in element:
        tag = i.tag.split('}')[1]
        print("\n",tag, xsd.keys())
        # for t in range(tabs):
        # row+=1
        # peas_of_sheat['A'+str(row)] = ""
        # print('  ',end='')
        if tag in xsd.keys():
            if 'annotation' in xsd[tag].keys():
                row += 1
                fill_one = PatternFill('solid', fgColor="556B2F")
                fill_two = PatternFill('solid', fgColor="9ACD32")
                
                peas_of_sheat['A'+str(row)].fill = fill_one
                peas_of_sheat['B'+str(row)].fill = fill_two

                peas_of_sheat['A'+str(row)].border = Border(top=Side(border_style="thin", 
                    color="808000"))
                peas_of_sheat['B'+str(row)].border = Border(top=Side(border_style="thin", 
                    color="808000"), left=Side(border_style="thin", color="808000"))

                peas_of_sheat['A'+str(row)] = xsd[tag]['annotation']
                peas_of_sheat['B'+str(row)] = i.text if not i.text is None and len(i.text)>0 else 'NONE'
                print("КОЛОНКИ", xsd[tag]['annotation'], i.text if not i.text is None and len(i.text)>0 else '')
            else:
                # row += 1
                # peas_of_sheat['A'+str(row)] = ""
                # peas_of_sheat['B'+str(row)] = i.text if not i.text is None and len(i.text)>0 else 'NONE'
                print("ХЗ", tag, i.text if not i.text is None and len(i.text)>0 else '')
            printXmlXsd(i, xsd[tag], tabs+1)
        else:
            # row += 1
            # peas_of_sheat['A'+str(row)] = ""

            row += 1
            fill_one = PatternFill('solid', fgColor="90EE90")
            fill_two = PatternFill('solid', fgColor="98FB98")

            peas_of_sheat['A'+str(row)].fill = fill_one
            peas_of_sheat['B'+str(row)].fill = fill_two
            
            peas_of_sheat['A'+str(row)].border = Border(top=Side(border_style="thin", 
                color="808000"))
            peas_of_sheat['B'+str(row)].border = Border(top=Side(border_style="thin", 
                color="808000"), left=Side(border_style="thin", color="808000"))

            peas_of_sheat['A'+str(row)] = tag
            peas_of_sheat['B'+str(row)] = i.text if not i.text is None and len(i.text)>0 else 'NONE'
            print("123", tag, i.text if not i.text is None and len(i.text)>0 else '')
            printXmlXsd(i, {}, tabs+1)

    # file = open(str_name, 'w')

    # for i in element:
    #     tag = i.tag.split('}')[1]
    #     for t in range(tabs):
    #         # file.write('  ')
    #         str_text += " "
    #         # print('  ', end='')
    #     if tag in xsd.keys():
    #         if 'annotation' in xsd[tag].keys():
    #             # file.write(str(xsd[tag]['annotation']))
    #             # file.write(str(i.text if not i.text is None and len(i.text)>0 else ''))
                
    #             str_text += str("\n") + str(xsd[tag]['annotation']) + str(" ") + str(i.text if not i.text is None and len(i.text)>0 else '')
                
    #             # print(xsd[tag]['annotation'], i.text if not i.text is None and len(i.text)>0 else '')
    #         else:
    #             # file.write(str(tag))
    #             # file.write(str(i.text if not i.text is None and len(i.text)>0 else ''))

    #             str_text += str("\n") + str(tag) + str(" ") + str(i.text if not i.text is None and len(i.text)>0 else '')

    #             # print(tag, i.text if not i.text is None and len(i.text)>0 else '')
    #         printXmlXsd(i, xsd[tag], tabs+1)
    #     else:
    #         # file.write(str(tag))
    #         # file.write(str(i.text if not i.text is None and len(i.text)>0 else ''))

    #         str_text += str("\n") + str(tag) + str(" ") + str(i.text if not i.text is None and len(i.text)>0 else '')
            
    #         # print(tag, i.text if not i.text is None and len(i.text)>0 else '')
    #         printXmlXsd(i, {}, tabs+1)

        
def main():
    app = QApplication(sys.argv)

    window = MainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == '__main__':
    main()